"""Excel export helpers for generated test packs."""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="171F33")
HEADER_FONT = Font(color="DAE2FD", bold=True)


def build_test_pack_workbook(payload: dict[str, Any]) -> bytes:
    """Build an XLSX workbook from a normalized generate response."""
    workbook = Workbook()
    cases_sheet = workbook.active
    cases_sheet.title = "Test Cases"

    headers = [
        "ID",
        "JIRA Ref",
        "Title",
        "Type",
        "Priority",
        "Objective",
        "Preconditions",
        "Test Data",
        "Steps",
        "Expected Result",
        "Mapped Acceptance Criteria",
        "Risk",
    ]
    cases_sheet.append(headers)

    for case in payload.get("test_cases", []) or []:
        cases_sheet.append(
            [
                case.get("id", ""),
                case.get("jira_ref", ""),
                case.get("title", ""),
                case.get("type", ""),
                case.get("priority", ""),
                case.get("objective", ""),
                _join(case.get("preconditions")),
                _join(case.get("test_data")),
                _join(case.get("steps")),
                case.get("expected_result", ""),
                _join(case.get("mapped_acceptance_criteria")),
                case.get("risk", ""),
            ]
        )

    _style_sheet(cases_sheet)

    traceability_sheet = workbook.create_sheet("Traceability")
    traceability_sheet.append(["Acceptance Criteria", "Test Case IDs", "Test Types", "Coverage Status"])
    for row in payload.get("traceability_matrix", []) or []:
        traceability_sheet.append(
            [
                row.get("acceptance_criteria", ""),
                row.get("test_case_ids", ""),
                row.get("test_types", ""),
                row.get("coverage_status", ""),
            ]
        )
    _style_sheet(traceability_sheet)

    coverage_sheet = workbook.create_sheet("Coverage Summary")
    coverage_sheet.append(["Coverage Item"])
    for item in payload.get("coverage_summary", []) or []:
        coverage_sheet.append([item])
    _style_sheet(coverage_sheet)

    summary_sheet = workbook.create_sheet("Summary")
    summary = payload.get("summary", {}) or {}
    summary_rows = [
        ("JIRA Key", payload.get("jira_key", "")),
        ("Status", payload.get("status", "")),
        ("Total Cases", summary.get("total_cases", 0)),
        ("JIRA Score", summary.get("jira_score", "")),
        ("RAG Score", summary.get("rag_score", "")),
        ("Source", summary.get("source", "")),
    ]
    for row in summary_rows:
        summary_sheet.append(row)
    summary_sheet.append(())
    summary_sheet.append(("Final Answer", payload.get("final_answer", "")))
    _style_sheet(summary_sheet)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _join(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        return "\n".join(str(item) for item in value)
    return str(value)


def _style_sheet(sheet) -> None:
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for column_cells in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(width + 2, 12), 52)
