from io import BytesIO

from openpyxl import load_workbook

from services.excel_export import build_test_pack_workbook
from services.response_normalizer import normalize_graph_result


def test_normalize_graph_result_flattens_specialist_outputs():
    result = normalize_graph_result(
        {
            "jira_key": "SCRUM-5",
            "jira_score": 0.8,
            "rag_score": 0.9,
            "jira_story": {"key": "SCRUM-5", "source": "mcp"},
            "is_safe_input": True,
            "is_safe_output": True,
            "functional_tests": [
                {
                    "id": "FT-001",
                    "title": "Verify valid login",
                    "type": "functional",
                    "priority": "High",
                    "steps": ["Open login", "Submit valid credentials"],
                    "expected_result": "User is signed in.",
                }
            ],
            "negative_tests": [
                {
                    "id": "NT-001",
                    "title": "Reject invalid password",
                    "steps": ["Open login", "Submit invalid password"],
                }
            ],
            "final_answer": "# Test Case Pack",
        }
    )

    assert result["status"] == "ready"
    assert result["summary"]["total_cases"] == 2
    assert result["summary"]["source"] == "mcp"
    assert result["test_cases"][0]["jira_ref"] == "SCRUM-5"
    assert result["test_cases"][1]["type"] == "negative"
    assert result["summary"]["score_mode"] == "live_response"


def test_normalize_graph_result_extracts_traceability_and_coverage_cards():
    result = normalize_graph_result(
        {
            "jira_key": "SCRUM-5",
            "jira_story": {"key": "SCRUM-5", "source": "mcp"},
            "is_safe_input": True,
            "is_safe_output": True,
            "final_answer": """
# Test Case Pack

## 4. Traceability Matrix

| Acceptance Criteria | Test Case IDs | Test Types | Coverage Status |
| --- | --- | --- | --- |
| User can log in | FT-001 | functional | Covered |

## 5. Coverage Summary

- Requested test types: functional
- Total test cases: 1
""",
        }
    )

    assert result["traceability_matrix"][0]["acceptance_criteria"] == "User can log in"
    assert result["coverage_summary"] == ["Requested test types: functional", "Total test cases: 1"]


def test_build_test_pack_workbook_contains_cases_and_summary():
    payload = {
        "jira_key": "SCRUM-5",
        "status": "ready",
        "summary": {"total_cases": 1, "jira_score": 0.8, "rag_score": 0.9, "source": "mcp"},
        "test_cases": [
            {
                "id": "FT-001",
                "jira_ref": "SCRUM-5",
                "title": "Verify valid login",
                "type": "functional",
                "priority": "High",
                "steps": ["Open login", "Submit valid credentials"],
                "expected_result": "User is signed in.",
            }
        ],
        "traceability_matrix": [
            {
                "acceptance_criteria": "User can log in",
                "test_case_ids": "FT-001",
                "test_types": "functional",
                "coverage_status": "Covered",
            }
        ],
        "coverage_summary": ["Requested test types: functional"],
        "final_answer": "# Test Case Pack",
    }

    workbook_bytes = build_test_pack_workbook(payload)
    workbook = load_workbook(BytesIO(workbook_bytes))

    assert workbook.sheetnames == ["Test Cases", "Traceability", "Coverage Summary", "Summary"]
    assert workbook["Test Cases"]["A2"].value == "FT-001"
    assert workbook["Traceability"]["A2"].value == "User can log in"
    assert workbook["Coverage Summary"]["A2"].value == "Requested test types: functional"
    assert workbook["Summary"]["A1"].value == "JIRA Key"
    assert workbook["Summary"]["B1"].value == "SCRUM-5"
